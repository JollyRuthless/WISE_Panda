"""
W.I.S.E. Panda — Great Hunt data helpers
Stores imported dropdown reference data and fight annotations.
"""

from __future__ import annotations

import csv
import json
import re
import sqlite3
import time as time_mod
from contextlib import closing
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from engine.parser_core import _open_log, parse_line


DATA_FILE = Path(__file__).parent.parent / "data" / "great_hunt_data.json"
DEFAULT_DB_FILE = Path(__file__).parent.parent / "data" / "great_hunt_data.sqlite3"
DB_TIMEOUT_SECONDS = 10
DB_LOCK_RETRY_DELAYS = (0.25, 0.5, 1.0)
ENTRY_TEXT_FIELDS = {
    "npc_entity_id",
    "mob_name",
    "classification",
    "location",
    "zone",
    "location_type",
    "instance_name",
    "quest_name",
    "character_name",
    "conflict",
    "last_zone_loaded",
    "abilities_used",
    "largest_hit_taken_by",
    "largest_hit_taken_ability",
    "largest_hit_done_target",
    "largest_hit_done_ability",
    "first_seen_date",
    "first_killed_date",
    "last_kill_date",
    "first_seen_by",
    "last_seen_by",
    "toughness",
    "journal_entry",
    "fight_notes",
    "picture_path",
}
ENTRY_INT_FIELDS = {
    "max_hp_seen",
    "mob_count",
    "kill_count",
    "total_damage_taken",
    "total_damage_done",
    "largest_hit_taken_amount",
    "largest_hit_done_amount",
}


def _default_payload() -> dict:
    return {
        "reference_rows": [],
        "annotations": {},
        "entries": {},
    }


def _db_path() -> Path:
    if DATA_FILE.name == "great_hunt_data.json":
        return DEFAULT_DB_FILE
    return DATA_FILE.with_suffix(".sqlite3")


def _connect_db() -> sqlite3.Connection:
    db_path = _db_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path), timeout=DB_TIMEOUT_SECONDS)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
    except sqlite3.OperationalError:
        pass
    return conn


def _run_db_operation(operation, context: str):
    last_error: Optional[sqlite3.OperationalError] = None
    for attempt, delay in enumerate((0.0, *DB_LOCK_RETRY_DELAYS), start=1):
        try:
            return operation()
        except sqlite3.OperationalError as exc:
            last_error = exc
            message = str(exc).lower()
            if "locked" not in message and "busy" not in message:
                raise RuntimeError(f"{context} failed: {exc}") from exc
            if attempt > len(DB_LOCK_RETRY_DELAYS):
                break
            time_mod.sleep(delay)
    raise RuntimeError(
        f"{context} failed because the Great Hunt database remained locked."
    ) from last_error


def _init_db() -> None:
    with closing(_connect_db()) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS great_hunt_reference_rows (
                kind TEXT NOT NULL,
                value TEXT NOT NULL,
                parent TEXT NOT NULL DEFAULT '',
                PRIMARY KEY (kind, value, parent)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS great_hunt_annotations (
                fight_key TEXT PRIMARY KEY,
                location_name TEXT NOT NULL DEFAULT '',
                zone_name TEXT NOT NULL DEFAULT '',
                location_type TEXT NOT NULL DEFAULT '',
                instance_name TEXT NOT NULL DEFAULT '',
                quest_name TEXT NOT NULL DEFAULT '',
                character_name TEXT NOT NULL DEFAULT '',
                fight_label TEXT NOT NULL DEFAULT '',
                log_path TEXT NOT NULL DEFAULT ''
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS great_hunt_annotation_mobs (
                fight_key TEXT NOT NULL,
                mob_key TEXT NOT NULL,
                mob_name TEXT NOT NULL DEFAULT '',
                npc_entity_id TEXT NOT NULL DEFAULT '',
                classification TEXT NOT NULL DEFAULT '',
                max_hp_seen INTEGER NOT NULL DEFAULT 0,
                instances_seen INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (fight_key, mob_key)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS great_hunt_entries (
                npc_entity_id TEXT PRIMARY KEY,
                mob_name TEXT NOT NULL DEFAULT '',
                classification TEXT NOT NULL DEFAULT '',
                location TEXT NOT NULL DEFAULT '',
                zone TEXT NOT NULL DEFAULT '',
                location_type TEXT NOT NULL DEFAULT '',
                instance_name TEXT NOT NULL DEFAULT '',
                quest_name TEXT NOT NULL DEFAULT '',
                character_name TEXT NOT NULL DEFAULT '',
                max_hp_seen INTEGER NOT NULL DEFAULT 0,
                mob_count INTEGER NOT NULL DEFAULT 0,
                conflict TEXT NOT NULL DEFAULT '',
                last_zone_loaded TEXT NOT NULL DEFAULT '',
                abilities_used TEXT NOT NULL DEFAULT '',
                kill_count INTEGER NOT NULL DEFAULT 0,
                total_damage_taken INTEGER NOT NULL DEFAULT 0,
                total_damage_done INTEGER NOT NULL DEFAULT 0,
                largest_hit_taken_amount INTEGER NOT NULL DEFAULT 0,
                largest_hit_taken_by TEXT NOT NULL DEFAULT '',
                largest_hit_taken_ability TEXT NOT NULL DEFAULT '',
                largest_hit_done_amount INTEGER NOT NULL DEFAULT 0,
                largest_hit_done_target TEXT NOT NULL DEFAULT '',
                largest_hit_done_ability TEXT NOT NULL DEFAULT '',
                first_seen_date TEXT NOT NULL DEFAULT '',
                first_killed_date TEXT NOT NULL DEFAULT '',
                last_kill_date TEXT NOT NULL DEFAULT '',
                first_seen_by TEXT NOT NULL DEFAULT '',
                last_seen_by TEXT NOT NULL DEFAULT '',
                toughness TEXT NOT NULL DEFAULT '',
                journal_entry TEXT NOT NULL DEFAULT '',
                fight_notes TEXT NOT NULL DEFAULT '',
                picture_path TEXT NOT NULL DEFAULT ''
            )
            """
        )
        _ensure_entry_columns(conn)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_hunt_annotation_mobs_npc_id ON great_hunt_annotation_mobs(npc_entity_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_hunt_entries_location ON great_hunt_entries(location, zone, location_type)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_hunt_annotations_location ON great_hunt_annotations(location_name, zone_name, location_type)")
        conn.commit()


def _ensure_entry_columns(conn: sqlite3.Connection) -> None:
    columns = {str(row[1]) for row in conn.execute("PRAGMA table_info(great_hunt_entries)").fetchall()}
    text_defaults = ENTRY_TEXT_FIELDS - columns
    int_defaults = ENTRY_INT_FIELDS - columns
    for column in sorted(text_defaults):
        conn.execute(f"ALTER TABLE great_hunt_entries ADD COLUMN {column} TEXT NOT NULL DEFAULT ''")
    for column in sorted(int_defaults):
        conn.execute(f"ALTER TABLE great_hunt_entries ADD COLUMN {column} INTEGER NOT NULL DEFAULT 0")


def _load_json_payload(path: Path) -> dict:
    try:
        if path.exists():
            raw = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                payload = _default_payload()
                payload.update(raw)
                return payload
    except Exception:
        pass
    return _default_payload()


def _db_has_rows() -> bool:
    db_path = _db_path()
    if not db_path.exists():
        return False
    with closing(_connect_db()) as conn:
        row = conn.execute("SELECT 1 FROM great_hunt_entries LIMIT 1").fetchone()
        if row is None:
            row = conn.execute("SELECT 1 FROM great_hunt_annotations LIMIT 1").fetchone()
        if row is None:
            row = conn.execute("SELECT 1 FROM great_hunt_reference_rows LIMIT 1").fetchone()
    return row is not None


def _migrate_json_if_needed() -> None:
    if _db_has_rows():
        return
    payload = _load_legacy_db_payload()
    if payload is None and DATA_FILE.exists():
        payload = _load_json_payload(DATA_FILE)
    if payload is None:
        return
    _write_payload_to_db(payload)


def load_data() -> dict:
    _init_db()
    _migrate_json_if_needed()
    try:
        return _load_payload_from_db()
    except Exception:
        legacy = _load_legacy_db_payload()
        if legacy is not None:
            return legacy
        return _load_json_payload(DATA_FILE)


def save_data(payload: dict) -> None:
    _init_db()
    normalized = _default_payload()
    normalized.update(payload or {})
    _write_payload_to_db(normalized)


def _load_payload_from_db() -> dict:
    payload = _default_payload()
    with closing(_connect_db()) as conn:
        conn.row_factory = sqlite3.Row

        payload["reference_rows"] = [
            {
                "kind": str(row["kind"] or "").strip(),
                "value": str(row["value"] or "").strip(),
                "parent": str(row["parent"] or "").strip(),
            }
            for row in conn.execute(
                """
                SELECT kind, value, parent
                FROM great_hunt_reference_rows
                ORDER BY kind, lower(parent), lower(value)
                """
            ).fetchall()
        ]

        annotations: dict[str, dict] = {}
        for row in conn.execute(
            """
            SELECT fight_key, location_name, zone_name, location_type, instance_name,
                   quest_name, character_name, fight_label, log_path
            FROM great_hunt_annotations
            ORDER BY fight_key
            """
        ).fetchall():
            fight_key = str(row["fight_key"] or "").strip()
            annotations[fight_key] = {
                "fight": {
                    "location_name": str(row["location_name"] or "").strip(),
                    "zone_name": str(row["zone_name"] or "").strip(),
                    "location_type": str(row["location_type"] or "").strip(),
                    "instance_name": str(row["instance_name"] or "").strip(),
                    "quest_name": str(row["quest_name"] or "").strip(),
                    "character_name": str(row["character_name"] or "").strip(),
                    "fight_label": str(row["fight_label"] or "").strip(),
                    "log_path": str(row["log_path"] or "").strip(),
                },
                "mobs": {},
            }
        for row in conn.execute(
            """
            SELECT fight_key, mob_key, mob_name, npc_entity_id, classification, max_hp_seen, instances_seen
            FROM great_hunt_annotation_mobs
            ORDER BY fight_key, mob_key
            """
        ).fetchall():
            fight_key = str(row["fight_key"] or "").strip()
            annotation = annotations.setdefault(fight_key, {"fight": {}, "mobs": {}})
            annotation["mobs"][str(row["mob_key"] or "").strip()] = {
                "mob_name": str(row["mob_name"] or "").strip(),
                "npc_entity_id": str(row["npc_entity_id"] or "").strip(),
                "classification": str(row["classification"] or "").strip(),
                "max_hp_seen": int(row["max_hp_seen"] or 0),
                "instances_seen": int(row["instances_seen"] or 0),
            }
        payload["annotations"] = annotations

        entries: dict[str, dict] = {}
        for row in conn.execute(
            """
            SELECT *
            FROM great_hunt_entries
            ORDER BY npc_entity_id
            """
        ).fetchall():
            npc_id = str(row["npc_entity_id"] or "").strip()
            entry = _empty_entry(npc_id)
            for field in ENTRY_TEXT_FIELDS:
                if field == "npc_entity_id":
                    continue
                entry[field] = str(row[field] or "").strip()
            for field in ENTRY_INT_FIELDS:
                entry[field] = str(int(row[field] or 0) or "")
            entries[npc_id] = entry
        payload["entries"] = entries
    return payload


def _write_payload_to_db(payload: dict) -> None:
    def _write() -> None:
        conn = _connect_db()
        try:
            conn.execute("DELETE FROM great_hunt_reference_rows")
            conn.execute("DELETE FROM great_hunt_annotations")
            conn.execute("DELETE FROM great_hunt_annotation_mobs")
            conn.execute("DELETE FROM great_hunt_entries")

            for row in payload.get("reference_rows", []):
                if not isinstance(row, dict):
                    continue
                conn.execute(
                    """
                    INSERT OR REPLACE INTO great_hunt_reference_rows (kind, value, parent)
                    VALUES (?, ?, ?)
                    """,
                    (
                        str(row.get("kind") or "").strip(),
                        str(row.get("value") or "").strip(),
                        str(row.get("parent") or "").strip(),
                    ),
                )

            annotations = payload.get("annotations", {})
            if isinstance(annotations, dict):
                for fight_key, annotation in annotations.items():
                    if not isinstance(annotation, dict):
                        continue
                    fight = annotation.get("fight", {})
                    if not isinstance(fight, dict):
                        fight = {}
                    conn.execute(
                        """
                        INSERT OR REPLACE INTO great_hunt_annotations (
                            fight_key, location_name, zone_name, location_type, instance_name,
                            quest_name, character_name, fight_label, log_path
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            str(fight_key or "").strip(),
                            str(fight.get("location_name") or fight.get("planet_name") or "").strip(),
                            str(fight.get("zone_name") or "").strip(),
                            str(fight.get("location_type") or "").strip(),
                            str(fight.get("instance_name") or "").strip(),
                            str(fight.get("quest_name") or "").strip(),
                            str(fight.get("character_name") or "").strip(),
                            str(fight.get("fight_label") or "").strip(),
                            str(fight.get("log_path") or "").strip(),
                        ),
                    )
                    mobs = annotation.get("mobs", {})
                    if not isinstance(mobs, dict):
                        continue
                    for mob_key, mob in mobs.items():
                        if not isinstance(mob, dict):
                            continue
                        conn.execute(
                            """
                            INSERT OR REPLACE INTO great_hunt_annotation_mobs (
                                fight_key, mob_key, mob_name, npc_entity_id, classification, max_hp_seen, instances_seen
                            ) VALUES (?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                str(fight_key or "").strip(),
                                str(mob_key or "").strip(),
                                str(mob.get("mob_name") or "").strip(),
                                str(mob.get("npc_entity_id") or "").strip(),
                                str(mob.get("classification") or "").strip(),
                                _safe_int(mob.get("max_hp_seen")),
                                _safe_int(mob.get("instances_seen")),
                            ),
                        )

            entries = payload.get("entries", {})
            if isinstance(entries, dict):
                for npc_id, entry in entries.items():
                    if not isinstance(entry, dict):
                        continue
                    normalized = _empty_entry(str(npc_id or entry.get("npc_entity_id") or "").strip())
                    normalized.update({key: str(value or "").strip() for key, value in entry.items() if key in normalized})
                    columns = [
                        "npc_entity_id", "mob_name", "classification", "location", "zone", "location_type",
                        "instance_name", "quest_name", "character_name", "max_hp_seen", "mob_count", "conflict",
                        "last_zone_loaded", "abilities_used", "kill_count", "total_damage_taken",
                        "total_damage_done", "largest_hit_taken_amount", "largest_hit_taken_by",
                        "largest_hit_taken_ability", "largest_hit_done_amount", "largest_hit_done_target",
                        "largest_hit_done_ability", "first_seen_date", "first_killed_date", "last_kill_date",
                        "first_seen_by", "last_seen_by", "toughness", "journal_entry", "fight_notes", "picture_path",
                    ]
                    placeholders = ", ".join("?" for _ in columns)
                    conn.execute(
                        f"""
                        INSERT OR REPLACE INTO great_hunt_entries ({", ".join(columns)})
                        VALUES ({placeholders})
                        """,
                        tuple(
                            _safe_int(normalized[field]) if field in ENTRY_INT_FIELDS else normalized[field]
                            for field in columns
                        ),
                    )
            conn.execute("DROP TABLE IF EXISTS great_hunt_store")
            conn.commit()
        finally:
            conn.close()

    _run_db_operation(_write, "Saving Great Hunt data")


def _load_legacy_db_payload() -> Optional[dict]:
    db_path = _db_path()
    if not db_path.exists():
        return None
    with closing(_connect_db()) as conn:
        tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        if "great_hunt_store" not in tables:
            return None
        rows = conn.execute("SELECT key, value_json FROM great_hunt_store").fetchall()
    payload = _default_payload()
    found = False
    for key, value_json in rows:
        if key not in payload:
            continue
        payload[key] = json.loads(value_json)
        found = True
    return payload if found else None


def import_reference_file(path: str) -> int:
    rows = _read_reference_rows(path)
    normalized: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for row in rows:
        kind = (
            row.get("kind")
            or row.get("type")
            or row.get("list_type")
            or row.get("category")
            or ""
        ).strip().lower()
        value = (
            row.get("value")
            or row.get("name")
            or row.get("label")
            or ""
        ).strip()
        parent = (
            row.get("parent")
            or row.get("group")
            or row.get("scope")
            or ""
        ).strip()
        if not kind or not value:
            continue
        item = (kind, value, parent)
        if item in seen:
            continue
        seen.add(item)
        normalized.append({
            "kind": kind,
            "value": value,
            "parent": parent,
        })

    payload = load_data()
    payload["reference_rows"] = sorted(
        normalized,
        key=lambda item: (item["kind"], item["parent"].lower(), item["value"].lower()),
    )
    save_data(payload)
    return len(normalized)


def get_choices(kind: str, parent: Optional[str] = None) -> list[str]:
    payload = load_data()
    rows = payload.get("reference_rows", [])
    values: list[str] = []
    for row in rows:
        if row.get("kind") != kind:
            continue
        row_parent = (row.get("parent") or "").strip()
        if parent:
            if row_parent and row_parent != parent:
                continue
        values.append(row.get("value", ""))
    return sorted({value for value in values if value}, key=str.lower)


def get_contextual_choices(
    kind: str,
    location: Optional[str] = None,
    zone: Optional[str] = None,
    location_type: Optional[str] = None,
) -> list[str]:
    """
    Return dropdown choices from imported reference rows plus saved entries.

    When location/zone/type context is provided, saved values from other
    locations are filtered out to keep the Great Hunt intake form focused.
    """
    wanted = kind.strip().lower()
    location = (location or "").strip()
    zone = (zone or "").strip()
    location_type = (location_type or "").strip()

    values: set[str] = set()
    if wanted in ("location", "planet"):
        values.update(get_choices("location"))
        values.update(get_choices("planet"))
    elif wanted == "zone":
        values.update(get_choices("zone", location or None))
    elif wanted == "instance":
        values.update(get_choices("instance", zone or location or None))
    elif wanted == "quest":
        values.update(get_choices("quest", zone or location or None))
    else:
        values.update(get_choices(wanted))

    payload = load_data()
    annotations = payload.get("annotations", {})
    if not isinstance(annotations, dict):
        return sorted({value for value in values if value}, key=str.lower)

    for annotation in annotations.values():
        if not isinstance(annotation, dict):
            continue
        fight = annotation.get("fight", {})
        if not isinstance(fight, dict):
            continue
        fight_location = (fight.get("location_name") or fight.get("planet_name") or "").strip()
        fight_zone = (fight.get("zone_name") or "").strip()
        fight_type = _stored_location_type(fight)
        if location and fight_location != location:
            continue
        if zone and fight_zone != zone:
            continue
        if location_type and fight_type != location_type:
            continue

        if wanted in ("location", "planet"):
            _add_nonblank(values, fight_location)
        elif wanted == "zone":
            _add_nonblank(values, fight_zone)
        elif wanted == "location_type":
            _add_nonblank(values, fight_type)
        elif wanted == "instance":
            _add_nonblank(values, _stored_instance_name(fight))
        elif wanted == "quest":
            _add_nonblank(values, (fight.get("quest_name") or "").strip())

    if not values and zone and wanted in ("location_type", "instance", "quest"):
        return get_contextual_choices(wanted, location=location, location_type=location_type)

    return sorted({value for value in values if value}, key=str.lower)


def get_recent_context_value(
    kind: str,
    location: Optional[str] = None,
    zone: Optional[str] = None,
    location_type: Optional[str] = None,
) -> str:
    """Return the most recently saved field value for the current context."""
    wanted = kind.strip().lower()
    location = (location or "").strip()
    zone = (zone or "").strip()
    location_type = (location_type or "").strip()

    payload = load_data()
    annotations = payload.get("annotations", {})
    if not isinstance(annotations, dict):
        return ""

    for annotation in reversed(list(annotations.values())):
        if not isinstance(annotation, dict):
            continue
        fight = annotation.get("fight", {})
        if not isinstance(fight, dict):
            continue
        fight_location = (fight.get("location_name") or fight.get("planet_name") or "").strip()
        fight_zone = (fight.get("zone_name") or "").strip()
        fight_type = _stored_location_type(fight)
        if location and fight_location != location:
            continue
        if zone and fight_zone != zone:
            continue
        if location_type and fight_type != location_type:
            continue

        if wanted in ("location", "planet"):
            value = fight_location
        elif wanted == "zone":
            value = fight_zone
        elif wanted == "location_type":
            value = fight_type
        elif wanted == "instance":
            value = _stored_instance_name(fight)
        elif wanted == "quest":
            value = (fight.get("quest_name") or "").strip()
        else:
            value = ""
        if value:
            return value
    return ""


def infer_location_fields(log_path: str, line_start: int = 0, line_end: Optional[int] = None) -> dict[str, str]:
    """
    Best-effort location inference for Great Hunt.
    Reads the most recent AreaEntered event up to the supplied line range and maps it
    onto imported location/zone/instance reference data when available.
    """
    area_name = _find_recent_area_name(log_path, line_start=line_start, line_end=line_end)
    if not area_name:
        return {}

    mapped = _map_area_to_reference(area_name)
    mapped["detected_area_name"] = area_name
    return mapped


def save_annotation(fight_key: str, annotation: dict) -> None:
    payload = load_data()
    existing_entries = _entry_map(payload, include_annotations=False)
    annotations = payload.setdefault("annotations", {})
    annotations[fight_key] = annotation
    entries: dict[str, dict[str, str]] = {}
    if isinstance(annotations, dict):
        for saved_annotation in annotations.values():
            _merge_annotation_into_entries(entries, saved_annotation)
    _overlay_manual_entry_fields(entries, existing_entries)
    payload["entries"] = entries
    save_data(payload)


def save_automatic_fight_data(fight_key: str, fight, location_fields: Optional[dict[str, str]] = None) -> None:
    """
    Record automatically discoverable mob data for a fight.

    This is intentionally independent of the Great Hunt review window so the
    app is always building the mob catalog while logs are opened or watched.
    """
    if fight is None:
        return
    try:
        fight.ensure_loaded()
    except Exception:
        return
    mob_payload = _automatic_mob_payload(fight)
    if not mob_payload:
        return

    existing = load_annotation(fight_key)
    existing_fight = existing.get("fight", {}) if isinstance(existing, dict) else {}
    existing_mobs = existing.get("mobs", {}) if isinstance(existing, dict) else {}
    location_fields = location_fields or {}
    fight_date = _fight_date_for(fight)
    fight_info = {
        "location_name": existing_fight.get("location_name") or location_fields.get("location_name", ""),
        "zone_name": existing_fight.get("zone_name") or location_fields.get("zone_name", ""),
        "location_type": existing_fight.get("location_type") or location_fields.get("location_type", ""),
        "instance_name": existing_fight.get("instance_name") or location_fields.get("instance_name", ""),
        "quest_name": existing_fight.get("quest_name", ""),
        "character_name": existing_fight.get("character_name") or (fight.player_name or ""),
        "fight_label": getattr(fight, "label", ""),
        "log_path": fight._log_path or "",
        "fight_date": existing_fight.get("fight_date") or fight_date,
    }

    payload = {"fight": fight_info, "mobs": {}}
    for mob_key, auto_mob in mob_payload.items():
        existing_mob = existing_mobs.get(mob_key, {}) if isinstance(existing_mobs, dict) else {}
        if not isinstance(existing_mob, dict):
            existing_mob = {}
        merged = dict(auto_mob)
        merged["classification"] = existing_mob.get("classification", "")
        for manual_field in ("mob_name",):
            if existing_mob.get(manual_field):
                merged[manual_field] = existing_mob[manual_field]
        payload["mobs"][mob_key] = merged
    save_annotation(fight_key, payload)


def load_annotation(fight_key: str) -> dict:
    payload = load_data()
    annotations = payload.get("annotations", {})
    existing = annotations.get(fight_key, {})
    return existing if isinstance(existing, dict) else {}


def clear_annotations() -> None:
    """Clear saved Great Hunt entries while preserving reference data."""
    payload = load_data()
    payload["annotations"] = {}
    payload["entries"] = {}
    save_data(payload)


def list_annotation_entries() -> list[dict[str, str]]:
    """Return one editable display row per NPC ID."""
    rows = list(_entry_map().values())
    return sorted(rows, key=_entry_sort_key)


def list_annotation_entry_page(search: str = "", limit: int = 100, offset: int = 0) -> list[dict[str, str]]:
    """Return one page of editable Great Hunt rows without loading the full table."""
    _init_db()
    _migrate_json_if_needed()
    search = str(search or "").strip().lower()
    limit = max(int(limit or 100), 1)
    offset = max(int(offset or 0), 0)
    params: list[object] = []
    where = ""
    if search:
        where = "WHERE lower(mob_name) LIKE ?"
        params.append(f"%{search}%")
    params.extend([limit, offset])
    with closing(_connect_db()) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            f"""
            SELECT *
            FROM great_hunt_entries
            {where}
            ORDER BY
                CASE
                    WHEN TRIM(classification) = '' OR lower(TRIM(classification)) = 'select'
                      OR TRIM(location) = '' OR TRIM(zone) = '' OR TRIM(location_type) = ''
                    THEN 0 ELSE 1
                END,
                CASE WHEN TRIM(quest_name) = '' THEN 1 ELSE 0 END,
                lower(quest_name),
                lower(mob_name),
                npc_entity_id
            LIMIT ? OFFSET ?
            """,
            tuple(params),
        ).fetchall()
    return [_entry_from_db_row(row) for row in rows]


def count_annotation_entries(search: str = "") -> dict[str, int]:
    """Return total/filter/conflict counts for the Great Hunt entries table."""
    _init_db()
    _migrate_json_if_needed()
    search = str(search or "").strip().lower()
    with closing(_connect_db()) as conn:
        total = int(conn.execute("SELECT COUNT(*) FROM great_hunt_entries").fetchone()[0] or 0)
        conflict_count = int(conn.execute(
            "SELECT COUNT(*) FROM great_hunt_entries WHERE TRIM(conflict) != ''"
        ).fetchone()[0] or 0)
        if search:
            filtered = int(conn.execute(
                "SELECT COUNT(*) FROM great_hunt_entries WHERE lower(mob_name) LIKE ?",
                (f"%{search}%",),
            ).fetchone()[0] or 0)
        else:
            filtered = total
    return {"total": total, "filtered": filtered, "conflicts": conflict_count}


def update_entry(npc_entity_id: str, updates: dict[str, str]) -> None:
    payload = load_data()
    entries = _entry_map(payload, include_annotations=True)
    npc_id = str(npc_entity_id or "").strip()
    if not npc_id:
        return
    entry = entries.get(npc_id, _empty_entry(npc_id))
    editable_fields = (ENTRY_TEXT_FIELDS | ENTRY_INT_FIELDS) - {"npc_entity_id"}
    for field, value in updates.items():
        if field not in editable_fields:
            continue
        text = str(value or "").strip()
        if field == "max_hp_seen":
            entry[field] = str(_safe_int(text) or "")
        else:
            entry[field] = text
    entry["conflict"] = ""
    entries[npc_id] = entry
    payload["entries"] = entries
    save_data(payload)


def import_missing_mobs_from_encounter_database(progress_callback=None, import_ids: Optional[Iterable[int]] = None) -> dict[str, int]:
    """Add and enrich Great Hunt NPCs from the imported combat-log database."""
    try:
        from storage.encounter_db import DB_PATH as ENCOUNTER_DB_PATH, init_db as init_encounter_db
    except Exception:
        return {"added": 0, "updated": 0, "processed": 0}

    init_encounter_db()
    if not ENCOUNTER_DB_PATH.exists():
        return {"added": 0, "updated": 0, "processed": 0}

    payload = load_data()
    entries = _entry_map(payload, include_annotations=True)
    summaries: dict[str, dict] = {}
    hp_by_instance: dict[str, int] = {}
    area_cache: dict[str, dict[str, str]] = {}
    current_import_id = None
    current_area = ""
    current_area_fields: dict[str, str] = {}
    processed = 0
    import_id_values = sorted({int(value) for value in (import_ids or []) if int(value) > 0})
    where_clause = ""
    query_params: list[int] = []
    if import_ids is not None:
        if not import_id_values:
            return {"added": 0, "updated": 0, "processed": 0}
        placeholders = ",".join("?" for _ in import_id_values)
        where_clause = f"WHERE e.import_id IN ({placeholders})"
        query_params = import_id_values

    def report(message: str) -> None:
        if progress_callback and progress_callback(processed, total_rows, message) is False:
            raise RuntimeError("Pull canceled.")

    with closing(sqlite3.connect(str(ENCOUNTER_DB_PATH))) as conn:
        conn.row_factory = sqlite3.Row
        total_rows = int(conn.execute(
            f"SELECT COUNT(*) FROM combat_log_events e {where_clause}",
            tuple(query_params),
        ).fetchone()[0] or 0)
        report("Scanning encounter database...")
        cursor = conn.execute(
            f"""
            SELECT
                e.import_id, e.line_number,
                e.source_name, e.source_type, e.source_entity_id, e.source_instance_id,
                e.source_hp, e.source_max_hp,
                e.target_name, e.target_type, e.target_entity_id, e.target_instance_id,
                e.target_hp, e.target_max_hp,
                e.ability_name, e.effect_type, e.effect_name, e.effect_detail_name,
                e.result_amount, e.result_type,
                i.log_path, i.source_character_name
            FROM combat_log_events e
            JOIN combat_log_imports i ON i.import_id = e.import_id
            {where_clause}
            ORDER BY e.import_id, e.line_number
            """,
            tuple(query_params),
        )
        while True:
            rows = cursor.fetchmany(50_000)
            if not rows:
                break
            for row in rows:
                processed += 1
                import_id = int(row["import_id"] or 0)
                if import_id != current_import_id:
                    current_import_id = import_id
                    current_area = ""
                    current_area_fields = {}

                if str(row["effect_type"] or "") == "AreaEntered" and str(row["effect_name"] or "").strip():
                    current_area = str(row["effect_name"] or "").strip()
                    if current_area not in area_cache:
                        area_cache[current_area] = _map_area_to_reference(current_area)
                    current_area_fields = area_cache[current_area]

                _merge_encounter_db_entity_summary(
                    summaries, row, "source", current_area, current_area_fields
                )
                _merge_encounter_db_entity_summary(
                    summaries, row, "target", current_area, current_area_fields
                )
                _merge_encounter_db_damage_summary(summaries, row)
                _merge_encounter_db_kill_summary(summaries, hp_by_instance, row)

            report(f"Scanning encounter database... {processed:,}/{total_rows:,}")

    added = 0
    updated = 0
    for npc_id, summary in summaries.items():
        existing = entries.get(npc_id, _empty_entry(npc_id))
        merged = _merge_imported_database_entry(existing, summary)
        if merged != existing:
            if npc_id not in entries:
                added += 1
            else:
                updated += 1
            entries[npc_id] = merged

    if not added and not updated:
        return {"added": 0, "updated": 0, "processed": processed}

    payload["entries"] = entries
    report("Saving Great Hunt table...")
    save_data(payload)
    return {"added": added, "updated": updated, "processed": processed}


def _stored_location_type(fight: dict) -> str:
    location_type = (fight.get("location_type") or "").strip()
    if location_type:
        return location_type
    instance_name = (fight.get("instance_name") or "").strip()
    if instance_name in ("Open World", "Instanced"):
        return instance_name
    return ""


def _entry_sort_key(row: dict[str, str]) -> tuple:
    quest_name = str(row.get("quest_name") or "").strip().lower()
    classification = str(row.get("classification") or "").strip()
    mob_name = str(row.get("mob_name") or "").strip().lower()
    npc_id = str(row.get("npc_entity_id") or "").strip()
    is_new = 0 if _entry_needs_review(row) else 1
    classification_rank = _classification_sort_rank(classification)
    quest_rank = (0, quest_name) if quest_name else (1, "")
    return (
        is_new,
        quest_rank,
        classification_rank,
        mob_name,
        npc_id,
    )


def _entry_needs_review(row: dict[str, str]) -> bool:
    return not (
        _classification_is_filled(str(row.get("classification") or ""))
        and str(row.get("location") or "").strip()
        and str(row.get("zone") or "").strip()
        and str(row.get("location_type") or "").strip()
    )


def _classification_sort_rank(value: str) -> int:
    cleaned = (value or "").strip().lower()
    ranks = {
        "normal": 0,
        "weak": 0,
        "normal (weak)": 0,
        "strong": 1,
        "silver": 1,
        "strong (silver)": 1,
        "elite": 2,
        "gold": 2,
        "elite (gold)": 2,
        "champion": 3,
        "boss": 4,
    }
    return ranks.get(cleaned, 99)


def _stored_instance_name(fight: dict) -> str:
    instance_name = (fight.get("instance_name") or "").strip()
    if instance_name in ("Open World", "Instanced"):
        return ""
    return instance_name


def _add_nonblank(values: set[str], value: str) -> None:
    if value:
        values.add(value)


def _display_values(values: set[str]) -> str:
    return " | ".join(sorted(values, key=str.lower))


def _safe_int(value) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _fight_date_for(fight) -> str:
    log_path = Path(fight._log_path) if getattr(fight, "_log_path", None) else None
    if log_path:
        path_date = _date_from_log_path(log_path)
        if path_date:
            return path_date
        try:
            return datetime.fromtimestamp(log_path.stat().st_mtime).date().isoformat()
        except OSError:
            pass
    return datetime.now().date().isoformat()


def _date_from_log_path(path: Path) -> str:
    match = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
    if match:
        return match.group(1)
    match = re.search(r"(\d{4})[-_](\d{2})[-_](\d{2})", path.name)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    return ""


def _automatic_mob_payload(fight) -> dict[str, dict]:
    fight_date = _fight_date_for(fight)
    mobs: dict[str, dict] = {}
    hp_by_instance: dict[str, int] = {}

    def is_player_controlled(entity) -> bool:
        return bool(getattr(entity, "player", None) or getattr(entity, "companion", None))

    def mob_key_for(entity) -> str:
        npc_id = getattr(entity, "npc_entity_id", "") or ""
        return f"{npc_id}|{getattr(entity, 'npc_instance', '') or ''}"

    def ensure_mob(entity) -> Optional[dict]:
        if not getattr(entity, "npc", None):
            return None
        npc_id = getattr(entity, "npc_entity_id", "") or ""
        if not npc_id:
            return None
        mob_key = mob_key_for(entity)
        row = mobs.setdefault(mob_key, {
            "mob_name": getattr(entity, "display_name", "") or getattr(entity, "npc", "") or "Unknown",
            "npc_entity_id": npc_id,
            "classification": "",
            "max_hp_seen": 0,
            "instances_seen": 0,
            "abilities": set(),
            "kill_count": 0,
            "total_damage_taken": 0,
            "total_damage_done": 0,
            "largest_hit_taken_amount": 0,
            "largest_hit_taken_by": "",
            "largest_hit_taken_ability": "",
            "largest_hit_done_amount": 0,
            "largest_hit_done_target": "",
            "largest_hit_done_ability": "",
            "first_seen_date": fight_date,
            "first_killed_date": "",
            "last_kill_date": "",
        })
        row["max_hp_seen"] = max(row["max_hp_seen"], getattr(entity, "maxhp", None) or 0)
        row["instances_seen"] = max(row["instances_seen"], 1)
        return row

    for ev in getattr(fight, "events", []) or []:
        src_mob = ensure_mob(ev.source)
        tgt_mob = ensure_mob(ev.target)
        ability_name = ev.ability.name if ev.ability else "Unknown"

        if src_mob and ability_name and ability_name != "Unknown":
            src_mob["abilities"].add(ability_name)

        if tgt_mob:
            instance_key = mob_key_for(ev.target)
            if getattr(ev.target, "hp", None) is not None:
                previous_hp = hp_by_instance.get(instance_key)
                if ev.target.hp <= 0 and (previous_hp is None or previous_hp > 0):
                    tgt_mob["kill_count"] += 1
                    tgt_mob["first_killed_date"] = tgt_mob["first_killed_date"] or fight_date
                    tgt_mob["last_kill_date"] = fight_date
                hp_by_instance[instance_key] = ev.target.hp

        if not (ev.is_damage and ev.result and not ev.result.is_miss):
            continue

        amount = int(ev.result.amount or 0)
        if tgt_mob and is_player_controlled(ev.source):
            tgt_mob["total_damage_taken"] += amount
            if amount > tgt_mob["largest_hit_taken_amount"]:
                tgt_mob["largest_hit_taken_amount"] = amount
                tgt_mob["largest_hit_taken_by"] = ev.source.display_name
                tgt_mob["largest_hit_taken_ability"] = ability_name

        if src_mob and is_player_controlled(ev.target):
            src_mob["total_damage_done"] += amount
            if amount > src_mob["largest_hit_done_amount"]:
                src_mob["largest_hit_done_amount"] = amount
                src_mob["largest_hit_done_target"] = ev.target.display_name
                src_mob["largest_hit_done_ability"] = ability_name

    output: dict[str, dict] = {}
    for mob_key, row in mobs.items():
        output[mob_key] = {
            "mob_name": row["mob_name"],
            "npc_entity_id": row["npc_entity_id"],
            "classification": "",
            "max_hp_seen": row["max_hp_seen"],
            "instances_seen": row["instances_seen"],
            "abilities_used": _display_values(row["abilities"]),
            "kill_count": row["kill_count"],
            "total_damage_taken": row["total_damage_taken"],
            "total_damage_done": row["total_damage_done"],
            "largest_hit_taken_amount": row["largest_hit_taken_amount"],
            "largest_hit_taken_by": row["largest_hit_taken_by"],
            "largest_hit_taken_ability": row["largest_hit_taken_ability"],
            "largest_hit_done_amount": row["largest_hit_done_amount"],
            "largest_hit_done_target": row["largest_hit_done_target"],
            "largest_hit_done_ability": row["largest_hit_done_ability"],
            "first_seen_date": row["first_seen_date"],
            "first_killed_date": row["first_killed_date"],
            "last_kill_date": row["last_kill_date"],
        }
    return output


def _latest_date(first: Optional[str], second: Optional[str]) -> str:
    first = (first or "").strip()
    second = (second or "").strip()
    if not first:
        return second
    if not second:
        return first
    return max(first, second)


def _merge_pipe_values(first: Optional[str], second: Optional[str]) -> str:
    values: set[str] = set()
    for raw in (first or "", second or ""):
        for part in raw.split("|"):
            cleaned = part.strip()
            if cleaned:
                values.add(cleaned)
    return _display_values(values)


def _overlay_manual_entry_fields(entries: dict[str, dict[str, str]], existing_entries: dict[str, dict[str, str]]) -> None:
    manual_fields = (
        "classification", "location", "zone", "location_type", "instance_name",
        "quest_name", "character_name", "toughness", "journal_entry",
        "fight_notes", "picture_path",
    )
    for npc_id, existing in existing_entries.items():
        if npc_id not in entries:
            entries[npc_id] = dict(existing)
            continue
        for field in manual_fields:
            if existing.get(field):
                entries[npc_id][field] = existing[field]


def has_complete_annotation(
    mob_keys: Iterable[str],
    location_fields: Optional[dict[str, str]] = None,
    fight_key: Optional[str] = None,
) -> bool:
    """
    Return True when Great Hunt already has a complete reusable entry.

    A completed entry has shared location info and a classification for every
    mob in the current encounter. If current location fields are known, they
    must match the saved shared fields so another planet/zone is not skipped.
    """
    keys = {key for key in mob_keys if key}
    if not keys:
        return False

    entry_map = _entry_map()
    location_fields = location_fields or {}
    for mob_key in keys:
        npc_id = mob_key.split("|", 1)[0].strip()
        entry = entry_map.get(npc_id)
        if not entry:
            return False
        if not _entry_has_shared_location(entry):
            return False
        if not _entry_location_matches(entry, location_fields):
            return False
        if not _classification_is_filled(entry.get("classification", "")):
            return False
    return True


def known_mob_classifications(
    mob_keys: Iterable[str],
    location_fields: Optional[dict[str, str]] = None,
    fight_key: Optional[str] = None,
) -> dict[str, str]:
    """
    Return saved classifications for any known mobs in the current encounter.

    Exact fight annotations are considered first; matching historical entries
    can fill blanks for mixed encounters where only some mobs are new.
    """
    keys = [key for key in mob_keys if key]
    if not keys:
        return {}

    known: dict[str, str] = {}
    entry_map = _entry_map()
    for mob_key in keys:
        npc_id = mob_key.split("|", 1)[0].strip()
        entry = entry_map.get(npc_id)
        if not entry:
            continue
        if not _entry_location_matches(entry, location_fields or {}):
            continue
        classification = (entry.get("classification") or "").strip()
        if classification:
            known[mob_key] = classification
    return known


def classification_for_npc(npc_entity_id: str) -> str:
    """Return the saved classification for a single NPC ID, if known."""
    npc_id = str(npc_entity_id or "").strip()
    if not npc_id:
        return ""
    entry = _entry_map().get(npc_id)
    if not entry:
        return ""
    classification = (entry.get("classification") or "").strip()
    if not _classification_is_filled(classification):
        return ""
    return classification


def _annotation_is_complete_for(annotation: dict, mob_keys: set[str], location_fields: dict[str, str]) -> bool:
    fight = annotation.get("fight", {})
    mobs = annotation.get("mobs", {})
    if not isinstance(fight, dict) or not isinstance(mobs, dict):
        return False
    if not _has_shared_location(fight):
        return False
    if not _location_matches(fight, location_fields):
        return False

    for mob_key in mob_keys:
        mob = mobs.get(mob_key, {})
        if not isinstance(mob, dict) or not _classification_is_filled(mob.get("classification", "")):
            return False
    return True


def _classification_is_filled(value: str) -> bool:
    cleaned = (value or "").strip()
    return bool(cleaned and cleaned.lower() != "select")


def _has_shared_location(fight: dict) -> bool:
    location = (fight.get("location_name") or fight.get("planet_name") or "").strip()
    return bool(
        location
        and (fight.get("zone_name") or "").strip()
        and _stored_location_type(fight)
    )


def _entry_has_shared_location(entry: dict) -> bool:
    return bool(
        (entry.get("location") or "").strip()
        and (entry.get("zone") or "").strip()
        and (entry.get("location_type") or "").strip()
    )


def _location_matches(saved: dict, current: dict[str, str]) -> bool:
    field_pairs = (
        ("location_name", ("location_name", "planet_name")),
        ("zone_name", ("zone_name",)),
        ("location_type", ("location_type",)),
        ("instance_name", ("instance_name",)),
    )
    for current_field, saved_fields in field_pairs:
        if current_field == "location_type":
            current_value = _stored_location_type(current)
        elif current_field == "instance_name":
            current_value = _stored_instance_name(current)
        else:
            current_value = (current.get(current_field) or "").strip()
        if not current_value:
            continue
        if current_field == "location_type":
            saved_value = _stored_location_type(saved)
        elif current_field == "instance_name":
            saved_value = _stored_instance_name(saved)
        else:
            saved_value = next(
                ((saved.get(field) or "").strip() for field in saved_fields if (saved.get(field) or "").strip()),
                "",
            )
        if saved_value != current_value:
            return False
    return True


def _entry_location_matches(saved: dict, current: dict[str, str]) -> bool:
    field_map = {
        "location_name": "location",
        "zone_name": "zone",
        "location_type": "location_type",
        "instance_name": "instance_name",
    }
    for current_field, saved_field in field_map.items():
        current_value = (
            _stored_location_type(current)
            if current_field == "location_type"
            else _stored_instance_name(current)
            if current_field == "instance_name"
            else (current.get(current_field) or current.get("planet_name") or "").strip()
        )
        if not current_value:
            continue
        saved_value = (saved.get(saved_field) or "").strip()
        if saved_value != current_value:
            return False
    return True


def _empty_entry(npc_id: str) -> dict[str, str]:
    return {
        "npc_entity_id": npc_id,
        "mob_name": "",
        "classification": "",
        "location": "",
        "zone": "",
        "location_type": "",
        "instance_name": "",
        "quest_name": "",
        "character_name": "",
        "max_hp_seen": "",
        "mob_count": "0",
        "conflict": "",
        "last_zone_loaded": "",
        "abilities_used": "",
        "kill_count": "0",
        "total_damage_taken": "0",
        "total_damage_done": "0",
        "largest_hit_taken_amount": "",
        "largest_hit_taken_by": "",
        "largest_hit_taken_ability": "",
        "largest_hit_done_amount": "",
        "largest_hit_done_target": "",
        "largest_hit_done_ability": "",
        "first_seen_date": "",
        "first_killed_date": "",
        "last_kill_date": "",
        "first_seen_by": "",
        "last_seen_by": "",
        "toughness": "",
        "journal_entry": "",
        "fight_notes": "",
        "picture_path": "",
    }


def _entry_from_db_row(row) -> dict[str, str]:
    npc_id = str(row["npc_entity_id"] or "").strip()
    entry = _empty_entry(npc_id)
    for field in ENTRY_TEXT_FIELDS:
        if field == "npc_entity_id":
            continue
        entry[field] = str(row[field] or "").strip()
    for field in ENTRY_INT_FIELDS:
        entry[field] = str(int(row[field] or 0) or "")
    return entry


def _merge_entry(existing: dict[str, str], incoming: dict[str, str]) -> dict[str, str]:
    merged = dict(existing)
    conflicts = {
        part.strip()
        for part in (merged.get("conflict") or "").split(",")
        if part.strip()
    }
    conflict_labels = {
        "mob_name": "Mob",
        "classification": "Type",
        "location": "Location",
        "zone": "Zone",
        "location_type": "Location Type",
        "instance_name": "Instance Name",
        "quest_name": "Quest Name",
    }
    for field in (
        "mob_name",
        "classification",
        "location",
        "zone",
        "location_type",
        "instance_name",
        "quest_name",
        "character_name",
        "last_zone_loaded",
        "first_seen_date",
        "first_killed_date",
        "first_seen_by",
    ):
        if (
            field in conflict_labels
            and merged.get(field)
            and incoming.get(field)
            and merged.get(field) != incoming.get(field)
        ):
            conflicts.add(conflict_labels[field])
        if not merged.get(field) and incoming.get(field):
            merged[field] = incoming[field]
    merged["max_hp_seen"] = str(max(_safe_int(merged.get("max_hp_seen")), _safe_int(incoming.get("max_hp_seen"))) or "")
    merged["mob_count"] = str(_safe_int(merged.get("mob_count")) + max(_safe_int(incoming.get("mob_count")), 1))
    merged["kill_count"] = str(_safe_int(merged.get("kill_count")) + _safe_int(incoming.get("kill_count")))
    merged["total_damage_taken"] = str(_safe_int(merged.get("total_damage_taken")) + _safe_int(incoming.get("total_damage_taken")))
    merged["total_damage_done"] = str(_safe_int(merged.get("total_damage_done")) + _safe_int(incoming.get("total_damage_done")))
    if _safe_int(incoming.get("largest_hit_taken_amount")) > _safe_int(merged.get("largest_hit_taken_amount")):
        merged["largest_hit_taken_amount"] = str(_safe_int(incoming.get("largest_hit_taken_amount")) or "")
        merged["largest_hit_taken_by"] = incoming.get("largest_hit_taken_by", "")
        merged["largest_hit_taken_ability"] = incoming.get("largest_hit_taken_ability", "")
    if _safe_int(incoming.get("largest_hit_done_amount")) > _safe_int(merged.get("largest_hit_done_amount")):
        merged["largest_hit_done_amount"] = str(_safe_int(incoming.get("largest_hit_done_amount")) or "")
        merged["largest_hit_done_target"] = incoming.get("largest_hit_done_target", "")
        merged["largest_hit_done_ability"] = incoming.get("largest_hit_done_ability", "")
    merged["last_kill_date"] = _latest_date(merged.get("last_kill_date"), incoming.get("last_kill_date"))
    merged["last_seen_by"] = incoming.get("last_seen_by") or merged.get("last_seen_by", "")
    merged["abilities_used"] = _merge_pipe_values(merged.get("abilities_used"), incoming.get("abilities_used"))
    merged["conflict"] = ", ".join(sorted(conflicts))
    return merged


def _merge_imported_database_entry(existing: dict[str, str], incoming: dict[str, str]) -> dict[str, str]:
    merged = dict(existing)
    for field in (
        "mob_name", "location", "zone", "location_type", "instance_name",
        "first_seen_date", "first_seen_by", "character_name", "last_zone_loaded",
    ):
        if not merged.get(field) and incoming.get(field):
            merged[field] = incoming[field]
    if incoming.get("last_seen_by"):
        merged["last_seen_by"] = incoming["last_seen_by"]
    merged["max_hp_seen"] = str(max(_safe_int(merged.get("max_hp_seen")), _safe_int(incoming.get("max_hp_seen"))) or "")
    merged["mob_count"] = str(max(_safe_int(merged.get("mob_count")), _safe_int(incoming.get("mob_count"))) or "0")
    for field in ("kill_count", "total_damage_taken", "total_damage_done"):
        merged[field] = str(max(_safe_int(merged.get(field)), _safe_int(incoming.get(field))) or "0")
    if _safe_int(incoming.get("largest_hit_taken_amount")) > _safe_int(merged.get("largest_hit_taken_amount")):
        merged["largest_hit_taken_amount"] = str(_safe_int(incoming.get("largest_hit_taken_amount")) or "")
        merged["largest_hit_taken_by"] = incoming.get("largest_hit_taken_by", "")
        merged["largest_hit_taken_ability"] = incoming.get("largest_hit_taken_ability", "")
    if _safe_int(incoming.get("largest_hit_done_amount")) > _safe_int(merged.get("largest_hit_done_amount")):
        merged["largest_hit_done_amount"] = str(_safe_int(incoming.get("largest_hit_done_amount")) or "")
        merged["largest_hit_done_target"] = incoming.get("largest_hit_done_target", "")
        merged["largest_hit_done_ability"] = incoming.get("largest_hit_done_ability", "")
    merged["abilities_used"] = _merge_pipe_values(merged.get("abilities_used"), incoming.get("abilities_used"))
    return merged


def _empty_encounter_db_summary(npc_id: str) -> dict:
    return {
        "npc_entity_id": npc_id,
        "mob_name": "",
        "location": "",
        "zone": "",
        "location_type": "",
        "instance_name": "",
        "last_zone_loaded": "",
        "character_name": "",
        "max_hp_seen": "0",
        "mob_count": "0",
        "kill_count": "0",
        "total_damage_taken": "0",
        "total_damage_done": "0",
        "largest_hit_taken_amount": "",
        "largest_hit_taken_by": "",
        "largest_hit_taken_ability": "",
        "largest_hit_done_amount": "",
        "largest_hit_done_target": "",
        "largest_hit_done_ability": "",
        "first_seen_date": "",
        "first_seen_by": "",
        "last_seen_by": "",
        "abilities_used": "",
        "_instances": set(),
        "_abilities": set(),
    }


def _summary_for_entity(summaries: dict[str, dict], npc_id: str) -> dict:
    return summaries.setdefault(npc_id, _empty_encounter_db_summary(npc_id))


def _merge_encounter_db_entity_summary(
    summaries: dict[str, dict],
    row,
    role: str,
    area_name: str,
    area_fields: dict[str, str],
) -> None:
    entity_type = str(row[f"{role}_type"] or "")
    if entity_type != "npc":
        return
    npc_id = str(row[f"{role}_entity_id"] or "").strip()
    if not npc_id:
        return
    summary = _summary_for_entity(summaries, npc_id)
    mob_name = str(row[f"{role}_name"] or "").strip()
    if not summary["mob_name"] and mob_name:
        summary["mob_name"] = mob_name
    summary["max_hp_seen"] = str(max(_safe_int(summary["max_hp_seen"]), _safe_int(row[f"{role}_max_hp"])))
    instance_id = str(row[f"{role}_instance_id"] or "").strip()
    if instance_id:
        summary["_instances"].add((int(row["import_id"] or 0), instance_id))
        summary["mob_count"] = str(len(summary["_instances"]))
    seen_by = str(row["source_character_name"] or "").strip()
    if not summary["first_seen_by"] and seen_by:
        summary["first_seen_by"] = seen_by
        summary["character_name"] = seen_by
    if seen_by:
        summary["last_seen_by"] = seen_by
    seen_date = _date_from_log_path(Path(str(row["log_path"] or "")))
    if seen_date and (not summary["first_seen_date"] or seen_date < summary["first_seen_date"]):
        summary["first_seen_date"] = seen_date
    if area_name and not summary["last_zone_loaded"]:
        summary["last_zone_loaded"] = area_name
    if area_fields:
        if not summary["location"]:
            summary["location"] = area_fields.get("location_name", "")
        if not summary["zone"]:
            summary["zone"] = area_fields.get("zone_name", "") or area_name
        if not summary["location_type"]:
            summary["location_type"] = area_fields.get("location_type", "")
        if not summary["instance_name"]:
            summary["instance_name"] = area_fields.get("instance_name", "")
    elif area_name and not summary["zone"]:
        summary["zone"] = area_name

    if role == "source":
        ability = str(row["ability_name"] or "").strip()
        if ability:
            summary["_abilities"].add(ability)
            summary["abilities_used"] = _display_values(summary["_abilities"])


def _merge_encounter_db_damage_summary(summaries: dict[str, dict], row) -> None:
    if str(row["effect_detail_name"] or "") != "Damage":
        return
    amount = _safe_int(row["result_amount"])
    if amount <= 0 or str(row["result_type"] or "").lower() in {"miss", "dodge", "parry", "deflect", "immune", "resist"}:
        return
    ability = str(row["ability_name"] or "").strip()
    source_type = str(row["source_type"] or "")
    target_type = str(row["target_type"] or "")

    if target_type == "npc" and source_type in {"player", "companion", "self"}:
        npc_id = str(row["target_entity_id"] or "").strip()
        if npc_id:
            summary = _summary_for_entity(summaries, npc_id)
            summary["total_damage_taken"] = str(_safe_int(summary["total_damage_taken"]) + amount)
            if amount > _safe_int(summary["largest_hit_taken_amount"]):
                summary["largest_hit_taken_amount"] = str(amount)
                summary["largest_hit_taken_by"] = str(row["source_name"] or "").strip()
                summary["largest_hit_taken_ability"] = ability

    if source_type == "npc" and target_type in {"player", "companion", "self"}:
        npc_id = str(row["source_entity_id"] or "").strip()
        if npc_id:
            summary = _summary_for_entity(summaries, npc_id)
            summary["total_damage_done"] = str(_safe_int(summary["total_damage_done"]) + amount)
            if amount > _safe_int(summary["largest_hit_done_amount"]):
                summary["largest_hit_done_amount"] = str(amount)
                summary["largest_hit_done_target"] = str(row["target_name"] or "").strip()
                summary["largest_hit_done_ability"] = ability


def _merge_encounter_db_kill_summary(summaries: dict[str, dict], hp_by_instance: dict[str, int], row) -> None:
    if str(row["target_type"] or "") != "npc":
        return
    npc_id = str(row["target_entity_id"] or "").strip()
    if not npc_id:
        return
    hp = row["target_hp"]
    if hp is None:
        return
    instance_key = f"{int(row['import_id'] or 0)}|{npc_id}|{row['target_instance_id'] or ''}"
    hp_value = _safe_int(hp)
    previous_hp = hp_by_instance.get(instance_key)
    if hp_value <= 0 and (previous_hp is None or previous_hp > 0):
        summary = _summary_for_entity(summaries, npc_id)
        summary["kill_count"] = str(_safe_int(summary["kill_count"]) + 1)
    hp_by_instance[instance_key] = hp_value


def _entry_from_annotation(fight: dict, mob: dict, mob_key: str) -> dict[str, str]:
    npc_id = str(mob.get("npc_entity_id") or mob_key.split("|", 1)[0] or "").strip()
    return {
        "npc_entity_id": npc_id,
        "mob_name": str(mob.get("mob_name") or mob_key.split("|")[0]).strip(),
        "classification": str(mob.get("classification") or "").strip(),
        "location": (fight.get("location_name") or fight.get("planet_name") or "").strip(),
        "zone": (fight.get("zone_name") or "").strip(),
        "location_type": _stored_location_type(fight),
        "instance_name": _stored_instance_name(fight),
        "quest_name": (fight.get("quest_name") or "").strip(),
        "character_name": (fight.get("character_name") or "").strip(),
        "max_hp_seen": str(_safe_int(mob.get("max_hp_seen")) or ""),
        "mob_count": str(max(_safe_int(mob.get("instances_seen")), 1)),
        "conflict": "",
        "last_zone_loaded": (fight.get("zone_name") or "").strip(),
        "abilities_used": str(mob.get("abilities_used") or "").strip(),
        "kill_count": str(_safe_int(mob.get("kill_count"))),
        "total_damage_taken": str(_safe_int(mob.get("total_damage_taken"))),
        "total_damage_done": str(_safe_int(mob.get("total_damage_done"))),
        "largest_hit_taken_amount": str(_safe_int(mob.get("largest_hit_taken_amount")) or ""),
        "largest_hit_taken_by": str(mob.get("largest_hit_taken_by") or "").strip(),
        "largest_hit_taken_ability": str(mob.get("largest_hit_taken_ability") or "").strip(),
        "largest_hit_done_amount": str(_safe_int(mob.get("largest_hit_done_amount")) or ""),
        "largest_hit_done_target": str(mob.get("largest_hit_done_target") or "").strip(),
        "largest_hit_done_ability": str(mob.get("largest_hit_done_ability") or "").strip(),
        "first_seen_date": str(mob.get("first_seen_date") or fight.get("fight_date") or "").strip(),
        "first_killed_date": str(mob.get("first_killed_date") or "").strip(),
        "last_kill_date": str(mob.get("last_kill_date") or "").strip(),
        "first_seen_by": (fight.get("character_name") or "").strip(),
        "last_seen_by": (fight.get("character_name") or "").strip(),
    }


def _entry_map(payload: Optional[dict] = None, *, include_annotations: bool = True) -> dict[str, dict[str, str]]:
    payload = payload or load_data()
    raw_entries = payload.get("entries", {})
    entries: dict[str, dict[str, str]] = {}
    if isinstance(raw_entries, dict):
        for npc_id, entry in raw_entries.items():
            if not isinstance(entry, dict):
                continue
            normalized = _empty_entry(str(npc_id or "").strip())
            normalized.update({key: str(value or "").strip() for key, value in entry.items() if key in normalized})
            if normalized["npc_entity_id"]:
                entries[normalized["npc_entity_id"]] = normalized
    if entries:
        return entries
    if not include_annotations:
        return entries

    annotations = payload.get("annotations", {})
    if not isinstance(annotations, dict):
        return entries
    for annotation in annotations.values():
        _merge_annotation_into_entries(entries, annotation)
    return entries


def _merge_annotation_into_entries(entries: dict[str, dict[str, str]], annotation: dict) -> None:
    if not isinstance(annotation, dict):
        return
    fight = annotation.get("fight", {})
    mobs = annotation.get("mobs", {})
    if not isinstance(fight, dict) or not isinstance(mobs, dict):
        return
    for mob_key, mob in mobs.items():
        if not isinstance(mob, dict):
            continue
        incoming = _entry_from_annotation(fight, mob, mob_key)
        npc_id = incoming["npc_entity_id"]
        if not npc_id:
            continue
        entries[npc_id] = _merge_entry(entries.get(npc_id, _empty_entry(npc_id)), incoming)


def _reference_rows() -> list[dict]:
    payload = load_data()
    rows = payload.get("reference_rows", [])
    return rows if isinstance(rows, list) else []


def _find_recent_area_name(log_path: str, line_start: int = 0, line_end: Optional[int] = None) -> Optional[str]:
    if not log_path:
        return None

    last_area: Optional[str] = None
    try:
        with _open_log(log_path) as handle:
            for idx, raw_line in enumerate(handle):
                if line_end is not None and idx > line_end:
                    break
                ev = parse_line(raw_line)
                if not ev:
                    continue
                if ev.effect_type == "AreaEntered" and ev.effect_name.strip():
                    last_area = ev.effect_name.strip()
    except Exception:
        return None
    return last_area


def _map_area_to_reference(area_name: str) -> dict[str, str]:
    area = area_name.strip()
    if not area:
        return {}

    locations = set(get_choices("location")) | set(get_choices("planet"))
    zones = set(get_choices("zone"))
    instances = set(get_choices("instance"))

    if area in locations:
        return {"location_name": area, "zone_name": "", "instance_name": ""}

    rows = _reference_rows()

    # Exact zone match. Infer parent location if the reference data provides it.
    for row in rows:
        if row.get("kind") == "zone" and (row.get("value") or "").strip() == area:
            location = (row.get("parent") or "").strip()
            return {"location_name": location, "zone_name": area, "instance_name": ""}

    # Exact instance match. Infer zone and maybe location via its parent row.
    for row in rows:
        if row.get("kind") == "instance" and (row.get("value") or "").strip() == area:
            zone = (row.get("parent") or "").strip()
            location = ""
            for parent_row in rows:
                if parent_row.get("kind") == "zone" and (parent_row.get("value") or "").strip() == zone:
                    location = (parent_row.get("parent") or "").strip()
                    break
            return {"location_name": location, "zone_name": zone, "instance_name": area}

    # If imported zone data already knows this area, keep it in zone.
    if area in zones:
        return {"location_name": "", "zone_name": area, "instance_name": ""}

    # Raw AreaEntered values are usually the top-level place the player entered.
    return {"location_name": area, "zone_name": "", "instance_name": ""}


def _read_reference_rows(path: str) -> list[dict[str, str]]:
    suffix = Path(path).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx_rows(path)
    return _read_csv_rows(path)


def _read_csv_rows(path: str) -> list[dict[str, str]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(2048)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        return [
            {str(key).strip().lower(): (value or "").strip() for key, value in row.items() if key}
            for row in reader
        ]


def _read_xlsx_rows(path: str) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(
            "Excel import requires openpyxl. Save the sheet as CSV, or install openpyxl."
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    output: list[dict[str, str]] = []
    for row in rows[1:]:
        item: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else ""
            item[header] = "" if value is None else str(value).strip()
        output.append(item)
    return output
